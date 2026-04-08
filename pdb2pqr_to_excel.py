"""
pdb2pqr_to_excel.py  —  Convert PDB2PQR stderr file(s) into an interactive
                         protein charge vs pH Excel workbook.

Supports two modes: single-file and batch.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SINGLE-FILE MODE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Process one PDB2PQR stderr file and produce one Excel workbook with a single
protein sheet.

Usage:
    python pdb2pqr_to_excel.py <stderr_file> [--output <output.xlsx>]
                               [--template <existing.xlsx>]
                               [--ph-step <step>] [--ph-override]

Arguments:
    stderr_file          PDB2PQR stderr output file containing a
                         'SUMMARY OF THIS PREDICTION' pKa table.
    --output <file>      Path for the output workbook.
                         Default: <stderr_file_stem>_pKa_Interactive.xlsx
                         saved alongside the input file.
    --template <file>    Optional existing .xlsx to copy additional sheets
                         from (e.g. your Pkas_Calculation.xlsx with other
                         protein sheets already in it).
    --ph-step <step>     pH axis step size. Default: 0.05.
                         Normal range: 0.05 to 1.0.
                         Steps below 0.05 require --ph-override.
    --ph-override        Unlock steps below 0.05 (down to 0.01).
                         Must be combined with --ph-step. Use with caution:
                         a step of 0.01 produces ~1400 rows per chain.

Example:
    python pdb2pqr_to_excel.py pdb2pqr_stderr_IsPETase.txt
    python pdb2pqr_to_excel.py pdb2pqr_stderr_IsPETase.txt --output IsPETase.xlsx
    python pdb2pqr_to_excel.py pdb2pqr_stderr_IsPETase.txt --ph-step 0.5
    python pdb2pqr_to_excel.py pdb2pqr_stderr_IsPETase.txt --ph-step 0.01 --ph-override  # fine resolution

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
BATCH MODE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Process all PDB2PQR stderr files in a directory and produce a single Excel
workbook with one protein sheet per file, plus shared Instructions and
AminoAcids sheets.

Usage:
    python pdb2pqr_to_excel.py --batch <directory> [--output <output.xlsx>]
                               [--pattern <glob>]
                               [--ph-step <step>] [--ph-override]

Arguments:
    --batch <directory>  Directory containing PDB2PQR stderr files.
    --output <file>      Path for the output workbook.
                         Default: pKa_Interactive_batch.xlsx inside the
                         batch directory.
    --pattern <glob>     Glob pattern to select files within the directory.
                         Default: *.txt
                         Example: --pattern "*stderr*.txt"
    --ph-step <step>     pH axis step size. Default: 0.05. See above.
    --ph-override        Unlock steps below 0.05. See above.

Example:
    python pdb2pqr_to_excel.py --batch ./stderr_files
    python pdb2pqr_to_excel.py --batch ./stderr_files --output batch.xlsx
    python pdb2pqr_to_excel.py --batch ./stderr_files --pattern "*stderr*.txt"
    python pdb2pqr_to_excel.py --batch ./stderr_files --ph-step 0.1

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
pH STEP SIZE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

The --ph-step flag controls the spacing of the pH axis (0.00 to 14.00):

    Step    pH rows    Notes
    1.0     15         Coarse; suitable for quick overview
    0.5     29         Moderate resolution
    0.1     141        Good balance of resolution and file size
    0.05    281        Default — fine resolution, manageable file size
    0.01    1401       High resolution; requires --ph-override; large files

Normal range is 0.05 to 1.0. Steps below 0.05 require --ph-override and are
blocked without it, because they produce significantly larger files
(e.g. 12 chains at step 0.01 = ~16,800 rows of formulas).
The absolute minimum step is 0.01.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SHEET NAMING
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Each protein sheet is named from the line in the stderr file that reads:
    INFO:Loading molecule: XXX.pdb
The sheet name will be XXX (up to 31 characters, Excel's sheet name limit).
If two files produce the same name a numeric suffix is appended (_1, _2, …).
If the line is not found the input filename stem is used as a fallback.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
WORKBOOK STRUCTURE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Every workbook (single or batch) contains:

  Instructions sheet  — Usage guide for the spreadsheet.
  <Protein> sheet(s)  — One sheet per protein. Each chain in a multi-chain
                        structure is a fully self-contained block stacked
                        vertically, separated by a labelled gap row.
      Col A       Notes (free text, no effect on calculation)
      Col B       Residue type
      Col C       Residue number
      Col D       Chain ID
      Col E       PROPKA pKa — edit this to override individual pKa values
      Col F       Model pKa (reference only, not used in calculation)
      Col G       Sign / Toggle — XLOOKUP from AminoAcids sheet (+1/-1/0)
      Col H       Notes (free text, no effect on calculation)
      Col I       pH axis (0.00 to 14.00, spacing set by --ph-step)
      Col J+      Henderson–Hasselbalch charge per residue at each pH
                  (one column per residue, number of columns varies by protein)
      Total       Sum of all active residue charges at each pH
      Rounded     Total charge rounded to nearest integer
      pH          pH value repeated next to the totals for easy reading
      Chart       Live charge vs pH line chart (one series per chain)
  AminoAcids sheet — Residue sign reference table (shared across all proteins).
                     Editing signs here propagates to all protein sheets via
                     the XLOOKUP formula in Col G.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SIGN CONVENTION AND TOGGLING
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Default signs (from AminoAcids sheet):
    ASP, GLU, CYS, TYR  →  -1  (acidic)
    HIS, LYS, ARG        →  +1  (basic)
    C-terminus (C-)      →   0  (excluded by default)
    N-terminus (N+)      →   0  (excluded by default)

To include C- and N+: set their charge to -1 and +1 respectively in the
AminoAcids sheet — all protein sheets update automatically.

Two ways to suppress a residue's contribution:
    Soft exclusion  — Set Col E (pKa) to 99.99. The residue remains visible
                      but its charge contribution is effectively 0 across all
                      physiological pH values.
    Hard exclusion  — Set Col G (Sign) to 0. The charge column returns
                      exactly 0 at all pH values.
"""

import re
import sys
import shutil
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ── Residue sign convention ───────────────────────────────────────────────────
SIGN_MAP = {
    'ASP': -1, 'GLU': -1, 'CYS': -1, 'TYR': -1, 'C-': -1,
    'HIS': +1, 'LYS': +1, 'ARG': +1, 'N+':  +1,
}

AA_TABLE = [
    ('Arg','Arginine',   +1,'Basic'),
    ('Lys','Lysine',     +1,'Basic'),
    ('His','Histidine',  +1,'Basic'),
    ('Asp','Aspartate',  -1,'Acidic'),
    ('Glu','Glutamate',  -1,'Acidic'),
    ('Ala','Alanine',     0,'Non-polar'),
    ('Asn','Asparagine',  0,'Polar'),
    ('Cys','Cysteine',   -1,'Acidic'),
    ('Gln','Glutamine',   0,'Polar'),
    ('Gly','Glycine',     0,'Non-polar'),
    ('Ile','Isoleucine',  0,'Non-polar'),
    ('Leu','Leucine',     0,'Non-polar'),
    ('Met','Methionine',  0,'Non-polar'),
    ('Phe','Phenylalanine',0,'Aromatic'),
    ('Pro','Proline',     0,'Non-polar'),
    ('Ser','Serine',      0,'Polar'),
    ('Thr','Threonine',   0,'Polar'),
    ('Trp','Tryptophan',  0,'Aromatic'),
    ('Tyr','Tyrosine',   -1,'Acidic'),
    ('Val','Valine',      0,'Non-polar'),
    ('C-', 'C-term',      0, None),
    ('N+', 'N-term',      0, None),
]

# ── Colours ───────────────────────────────────────────────────────────────────
DARK='1F3864'; MID='2E75B6'; LIGHT='D6E4F0'; WHITE='FFFFFF'
YELLOW='FFF2CC'; GREEN='E2EFDA'; GREY='F2F2F2'; RED_L='FCE4D6'

# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(c):
    return PatternFill("solid", fgColor=c)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

# ── Step 1: parse PDB2PQR file ────────────────────────────────────────────────
def parse_pdb2pqr(input_path: Path) -> list[dict]:
    """
    Extract the SUMMARY OF THIS PREDICTION table from a PDB2PQR stderr file.
    Returns a list of dicts with keys:
        residue_type, residue_number, chain, pKa, model_pKa, sign
    """
    row_pattern = re.compile(
        r"^\s+"
        r"([A-Z][A-Z0-9+\-]*)"   # residue type  (ASP, N+, C-, …)
        r"\s+(\d+)"               # residue number
        r"\s+([A-Z])"             # chain ID
        r"\s+(-?\d+(?:\.\d+)?)"  # pKa
        r"\s+(-?\d+(?:\.\d+)?)"  # model-pKa
        r"\s*$"
    )

    in_summary = False
    records = []

    with open(input_path, encoding="utf-8") as fh:
        for line in fh:
            if "SUMMARY OF THIS PREDICTION" in line:
                in_summary = True
                continue
            if not in_summary:
                continue
            stripped = line.strip()
            if not stripped or stripped.startswith("Group"):
                continue
            if re.match(r"(INFO|WARNING|ERROR):", stripped):
                break
            m = row_pattern.match(line)
            if m:
                rtype, rnum, chain, pka, model_pka = m.groups()
                sign = SIGN_MAP.get(rtype.upper(), 0)
                records.append({
                    "residue_type":   rtype,
                    "residue_number": int(rnum),
                    "chain":          chain,
                    "pKa":            float(pka),
                    "model_pKa":      float(model_pka),
                    "sign":           sign,
                })

    if not records:
        raise ValueError(
            f"No pKa summary found in '{input_path}'.\n"
            "Ensure the file contains a 'SUMMARY OF THIS PREDICTION' section."
        )

    return records

# ── Step 2: build Excel sheets ────────────────────────────────────────────────
def build_aa_sheet(wb):
    if 'AminoAcids' in wb.sheetnames:
        del wb['AminoAcids']
    ws = wb.create_sheet('AminoAcids')
    for c, h in enumerate(
        ['Three-Letter (Key)', 'Full Name', 'Net Charge (pH 7.4)', 'Class'], 1
    ):
        cell = ws.cell(1, c, h)
        cell.font = _font(True, WHITE)
        cell.fill = _fill(DARK)
        cell.alignment = _center()
        cell.border = _border()

    for r, (code, full, charge, cls) in enumerate(AA_TABLE, 2):
        bg = GREEN if charge > 0 else (RED_L if charge < 0 else GREY)
        for c, val in enumerate([code, full, charge, cls or ''], 1):
            cell = ws.cell(r, c, val)
            cell.font = _font()
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.alignment = _center() if c > 2 else _left()

    for col, w in zip('ABCD', [18, 18, 20, 14]):
        ws.column_dimensions[col].width = w


def build_instructions(wb, protein_name: str):
    if 'Instructions' in wb.sheetnames:
        del wb['Instructions']
    ws = wb.create_sheet('Instructions')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 72

    ws.merge_cells('B2:C2')
    c = ws['B2']
    c.value = "How to Use This Workbook"
    c.font = _font(True, WHITE, 14)
    c.fill = _fill(DARK)
    c.alignment = _center()

    sections = [
        ("COLUMN LAYOUT", None),
        ("Col A", "Notes — free text, no effect on calculation"),
        ("Col B", "Residue type (ASP, GLU, HIS, CYS, TYR, LYS, ARG, C-, N+)"),
        ("Col C", "Residue number"),
        ("Col D", "Chain ID"),
        ("Col E", "► PROPKA-calculated pKa  ◄  Edit this to override individual pKa values"),
        ("Col F", "Reference model pKa (display only — not used in calculation)"),
        ("Col G", "Charge sign  |  +1 = basic  |  -1 = acidic  |  0 = EXCLUDED from sum"),
        ("Col H", "Notes — free text, no effect on calculation"),
        ("Col I", "pH axis (0.00 to 14.00 in 0.01 steps), rows 2 onwards"),
        ("Col J+", "Henderson-Hasselbalch charge per residue at each pH (one column per residue)"),
        ("Total Charge", "Sum of all active (sign not 0) residue charges at each pH"),
        ("Rounded", "Total charge rounded to nearest integer"),
        ("", "Cols A-G and the pH/charge table share the same rows — residue data"),
        ("", "fills rows 2 to n+1, pH rows 2 to 1402, all side by side from row 1."),
        (None, None),
        ("HOW TO EXCLUDE A RESIDUE", None),
        ("Step 1", "Find the residue's row in the top section of the sheet"),
        ("Step 2", "Set Column G (Sign / Toggle) to  0"),
        ("Result", "That residue is removed from the total charge sum immediately"),
        ("To re-include", "Restore Column G to  +1  (basic) or  -1  (acidic)"),
        (None, None),
        ("HOW TO OVERRIDE A pKa", None),
        ("Step 1", "Find the residue's row"),
        ("Step 2", "Edit the value in Column E"),
        ("Result", "Charge curve and total charge update automatically across all pH rows"),
        (None, None),
        ("NOTES ON CYS & TYR", None),
        ("Sign = -1", "Both CYS and TYR are acidic residues with sign = -1 by default."),
        ("pKa = 99.99", "A pKa of 99.99 means the residue never deprotonates at physiological pH"),
        ("", "— its charge contribution is effectively 0 without changing the sign."),
        ("Method 1", "Set Column E (pKa) to 99.99 to silently suppress a residue's contribution"),
        ("", "while keeping it visible in the calculation (soft exclusion)."),
        ("Method 2", "Set Column G (Sign) to 0 to fully exclude a residue from the sum"),
        ("", "(hard exclusion — charge column returns 0 at all pH values)."),
        (None, None),
        ("TERMINI (C- and N+)", None),
        ("Default = 0", "The C-terminus (C-) and N-terminus (N+) are set to charge 0 by default"),
        ("", "in the AminoAcids sheet, so they contribute nothing to the total charge."),
        ("To include", "In the AminoAcids sheet, set C- to -1 and/or N+ to +1. Column G"),
        ("", "will update automatically via the XLOOKUP formula for all protein sheets."),
        ("", "Alternatively, override Column G directly for a specific residue row."),
        (None, None),
        ("CHART", None),
        ("", f"The chart on the '{protein_name}' sheet updates automatically when you"),
        ("", "change pKa values or toggle residues in Column G."),
        (None, None),
        ("RUNNING THE SCRIPT AGAIN", None),
        ("Command", "python pdb2pqr_to_excel.py  <your_pdb2pqr_file.txt>  [output.xlsx]"),
        ("Template", "Add  --template <existing.xlsx>  to carry over other protein sheets"),
    ]

    row = 4
    for key, val in sections:
        if key is None:
            row += 1
            continue
        if val is None:
            ws.merge_cells(f'B{row}:C{row}')
            c = ws.cell(row, 2, key)
            c.font = _font(True, WHITE)
            c.fill = _fill(MID)
            c.alignment = _left()
        else:
            c = ws.cell(row, 2, key)
            c.font = _font(True)
            c.fill = _fill(LIGHT)
            c.alignment = _left()
            c = ws.cell(row, 3, val)
            c.font = _font()
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 18
        row += 1


def build_protein_sheet(wb, sheet_name: str, records: list[dict], ph_step: float = 0.05):
    """
    Build one protein sheet. Each chain is a fully self-contained horizontal
    block stacked vertically, separated by a single labelled gap row.

    Each block contains:
      Col A        : Notes (free)
      Col B        : Residue type
      Col C        : Residue number
      Col D        : Chain ID
      Col E        : PROPKA pKa (editable, yellow)
      Col F        : Model pKa (reference)
      Col G        : Sign / Toggle (XLOOKUP from AminoAcids, yellow)
      Col H        : Notes (free)
      Col I        : pH axis (0.00 → 14.00)
      Col J+       : One Henderson–Hasselbalch charge column per residue
      Total col    : SUM of all active charges at each pH
      Rounded col  : Total rounded to nearest integer
      pH2 col      : pH repeated for easy reading

    Row 1 of each block is the header row for that block.
    Rows 2..1402 of each block are the pH rows.
    Each block is therefore n_ph + 1 rows tall (header + 1401 pH steps).
    Chains are separated by 1 blank gap row.
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    ph_values = [round(i * ph_step, 2) for i in range(int(round(14.0 / ph_step)) + 1)]
    n_ph = len(ph_values)

    # Fixed column indices (same for every block)
    COL_NOTES1 = 1   # A
    COL_RTYPE  = 2   # B
    COL_RNUM   = 3   # C
    COL_CHAIN  = 4   # D
    COL_PKA    = 5   # E
    COL_MPKA   = 6   # F
    COL_SIGN   = 7   # G
    COL_NOTES2 = 8   # H
    COL_PH     = 9   # I
    COL_RES0   = 10  # J — first charge column (same offset in every block)

    # Split records by chain, preserving order of first appearance
    from collections import defaultdict
    from openpyxl.chart.series import SeriesLabel
    chains_seen = []
    chains_dict = defaultdict(list)
    for rec in records:
        ch = rec['chain']
        if ch not in chains_dict:
            chains_seen.append(ch)
        chains_dict[ch].append(rec)
    chains = chains_seen

    # Block height = header + whichever is taller: pH rows or longest chain.
    # Matters when ph_step is large (e.g. 1.0 → 15 pH rows) but chains are long.
    max_chain_len = max(len(chains_dict[ch]) for ch in chains)
    block_height  = 1 + max(n_ph, max_chain_len)   # 1 header row + data rows

    # ── Calculate the block start row for each chain ──────────────────────────
    # Block i starts at: 1 + i * (block_height + 1)   [+1 for the gap row]
    # Chain 0: row 1, Chain 1: row block_height+2, etc.
    def block_start(i):
        return 1 + i * (block_height + 1)

    # ── Helper: write one complete chain block ────────────────────────────────
    def write_block(ch, block_row):
        recs_ch = chains_dict[ch]
        n_ch    = len(recs_ch)
        COL_TOTAL = COL_RES0 + n_ch
        COL_ROUND = COL_TOTAL + 1
        COL_PH2   = COL_ROUND + 1

        hdr_row = block_row   # header row for this block

        # ── Header row ────────────────────────────────────────────────────────
        static_hdrs = {
            COL_NOTES1: ('Notes',          GREY, False),
            COL_RTYPE:  ('Res Type',       DARK, True),
            COL_RNUM:   ('Res #',          DARK, True),
            COL_CHAIN:  ('Chain',          DARK, True),
            COL_PKA:    ('pKa\n(edit me)', DARK, True),
            COL_MPKA:   ('Model pKa',      DARK, True),
            COL_SIGN:   ('Sign /\nToggle', DARK, True),
            COL_NOTES2: ('Notes',          GREY, False),
            COL_PH:     ('pH',             DARK, True),
            COL_TOTAL:  ('Total Charge',   DARK, True),
            COL_ROUND:  ('Rounded',        MID,  True),
            COL_PH2:    ('pH',             DARK, True),
        }
        for col, (label, bg, bold) in static_hdrs.items():
            cell = ws.cell(hdr_row, col, label)
            cell.font      = _font(bold, WHITE, 9)
            cell.fill      = _fill(bg)
            cell.alignment = _center()
            cell.border    = _border()

        # Residue name headers (one per charge column)
        for j, rec in enumerate(recs_ch):
            cell = ws.cell(hdr_row, COL_RES0 + j,
                           f"{rec['residue_type']}{rec['residue_number']}")
            cell.font      = _font(False, '404040', 7)
            cell.fill      = _fill(LIGHT)
            cell.alignment = _center()
            cell.border    = _border()

        ws.row_dimensions[hdr_row].height = 30

        # ── Residue metadata rows (cols A–H, rows hdr+1 .. hdr+n_ch) ─────────
        for i, rec in enumerate(recs_ch):
            r = hdr_row + 1 + i
            row_data = [
                (COL_NOTES1, None,                  GREY,   False),
                (COL_RTYPE,  rec['residue_type'],   LIGHT,  True),
                (COL_RNUM,   rec['residue_number'], GREY,   False),
                (COL_CHAIN,  rec['chain'],           GREY,   False),
                (COL_PKA,    rec['pKa'],             YELLOW, True),
                (COL_MPKA,   rec['model_pKa'],       GREY,   False),
                (COL_NOTES2, None,                  GREY,   False),
            ]
            for col, val, bg, bold in row_data:
                cell           = ws.cell(r, col, val)
                cell.font      = _font(bold, size=10)
                cell.fill      = _fill(bg)
                cell.border    = _border()
                cell.alignment = _center()

            sign_cell = ws.cell(r, COL_SIGN,
                f'=_xlfn.XLOOKUP(PROPER(B{r}),AminoAcids!$A$2:$A$23,AminoAcids!$C$2:$C$23,0)')
            sign_cell.font          = _font(True, size=10)
            sign_cell.fill          = _fill(YELLOW)
            sign_cell.border        = _border()
            sign_cell.alignment     = _center()
            sign_cell.number_format = '0'
            ws.row_dimensions[r].height = 16

        # ── pH + charge rows ─────────────────────────────────────────────────
        # pH rows start at hdr_row + 1, sharing row space with residue metadata.
        # Residue metadata only fills cols A–H; pH/charge fills cols I+.
        # For pH rows beyond the residue count, cols A–H are blank.
        first_ch  = get_column_letter(COL_RES0)
        last_ch   = get_column_letter(COL_RES0 + n_ch - 1)
        total_ltr = get_column_letter(COL_TOTAL)

        for idx, ph in enumerate(ph_values):
            r      = hdr_row + 1 + idx
            stripe = GREY if idx % 2 == 0 else WHITE

            # pH value
            cell               = ws.cell(r, COL_PH, ph)
            cell.number_format = '0.00'
            cell.font          = _font(size=9)
            cell.fill          = _fill(stripe)
            cell.border        = _border()

            # Charge per residue — reference the residue metadata rows
            # which are at hdr_row+1 .. hdr_row+n_ch (absolute)
            for j in range(n_ch):
                res_row = hdr_row + 1 + j
                col     = COL_RES0 + j
                formula = f'=$G${res_row}*1/(1+10^(($I{r}-$E${res_row})*$G${res_row}))'
                cell               = ws.cell(r, col, formula)
                cell.number_format = '0.000'
                cell.font          = _font(size=8, color='404040')
                cell.fill          = _fill(stripe)
                cell.border        = _border()

            # Total
            cell               = ws.cell(r, COL_TOTAL,
                                         f'=SUM({first_ch}{r}:{last_ch}{r})')
            cell.number_format = '0.00'
            cell.font          = _font(True, size=9)
            cell.fill          = _fill(LIGHT)
            cell.border        = _border()

            # Rounded
            cell               = ws.cell(r, COL_ROUND,
                                         f'=ROUND({total_ltr}{r},0)')
            cell.number_format = '0'
            cell.font          = _font(True, size=9)
            cell.fill          = _fill(GREEN)
            cell.border        = _border()

            # pH repeated
            cell               = ws.cell(r, COL_PH2, f'=$I{r}')
            cell.number_format = '0.00'
            cell.font          = _font(True, size=9)
            cell.fill          = _fill(LIGHT)
            cell.border        = _border()

        return COL_TOTAL, COL_ROUND, COL_PH2

    # ── Write all chain blocks ────────────────────────────────────────────────
    chain_total_cols = {}   # for charting
    chain_ph_rows    = {}   # for charting

    for ci, ch in enumerate(chains):
        brow = block_start(ci)

        # Gap row label above every block except the first
        if ci > 0:
            gap_row = brow - 1
            cell = ws.cell(gap_row, COL_NOTES1, f'── Chain {ch} ──')
            cell.font      = _font(True, DARK, 10)
            cell.alignment = _left()

        total_col, round_col, ph2_col = write_block(ch, brow)
        chain_total_cols[ch] = total_col
        chain_ph_rows[ch]    = (brow + 1, brow + n_ph)   # first and last pH row

    # ── Column widths (same for all blocks since columns are shared) ──────────
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 6

    # Charge columns: use the widest chain to set column widths
    max_n = max(len(chains_dict[ch]) for ch in chains)
    for j in range(max_n):
        ws.column_dimensions[get_column_letter(COL_RES0 + j)].width = 8
    # Total, Rounded, pH2 — position depends on largest chain
    max_total = COL_RES0 + max_n
    ws.column_dimensions[get_column_letter(max_total)].width     = 12
    ws.column_dimensions[get_column_letter(max_total + 1)].width = 9
    ws.column_dimensions[get_column_letter(max_total + 2)].width = 6

    # ── Chart: one series per chain, each using its own pH and Total rows ─────
    COLOURS = ['1F3864','C00000','375623','7030A0','833C00','006B6B',
               'BF9000','4472C4','ED7D31','A9D18E','FF0000','00B0F0']

    chart              = LineChart()
    chart.title        = f"{sheet_name} – Net Charge vs pH"
    chart.style        = 10
    chart.y_axis.title = "Net Charge"
    chart.x_axis.title = "pH"
    chart.height       = 14
    chart.width        = 28

    for ci, ch in enumerate(chains):
        n_ch      = len(chains_dict[ch])
        total_col = COL_RES0 + n_ch   # same offset in every block
        ph_r1, ph_r2 = chain_ph_rows[ch]

        y_ref = Reference(ws, min_col=total_col, min_row=ph_r1, max_row=ph_r2)
        x_ref = Reference(ws, min_col=COL_PH,    min_row=ph_r1, max_row=ph_r2)
        chart.add_data(y_ref, titles_from_data=False)
        series = chart.series[ci]
        series.graphicalProperties.line.solidFill = COLOURS[ci % len(COLOURS)]
        series.graphicalProperties.line.width     = 20000
        series.title = SeriesLabel(v=f"Chain {ch}")
        if ci == 0:
            chart.set_categories(x_ref)

    anchor_col = get_column_letter(max_total + 4)
    ws.add_chart(chart, f"{anchor_col}1")

    return ws


def parse_molecule_name(input_path: Path) -> str:
    """
    Extract the sheet name from the 'Loading molecule: XXX.pdb' line.
    Falls back to the file stem if the line is not found.
    Returns a cleaned string safe for use as an Excel sheet name.
    """
    pattern = re.compile(r'Loading molecule:\s*(.+?)\.pdb', re.IGNORECASE)
    try:
        with open(input_path, encoding='utf-8') as fh:
            for line in fh:
                m = pattern.search(line)
                if m:
                    name = m.group(1).strip()
                    # Strip any leading path components (safety)
                    name = Path(name).name
                    # Sanitise for Excel: max 31 chars, no special chars
                    name = re.sub(r'[\\/*?:\[\]]', '_', name)[:31]
                    return name
    except Exception:
        pass
    # Fallback: use file stem
    return re.sub(r'[\\/*?:\[\]]', '_', input_path.stem)[:31]


def batch_main(input_dir: Path, output_path: Path, pattern: str, ph_step: float = 0.05) -> None:
    """
    Process all PDB2PQR stderr files matching `pattern` in `input_dir`,
    writing one Excel workbook with:
      - Instructions sheet
      - AminoAcids sheet
      - One protein sheet per input file, named from 'Loading molecule: XXX.pdb'
    """
    files = sorted(input_dir.glob(pattern))
    if not files:
        sys.exit(f"No files matching '{pattern}' found in {input_dir}")

    print(f"Found {len(files)} file(s) in {input_dir}")

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    build_aa_sheet(wb)
    # Build instructions with a generic name; we'll update after we know all sheets
    build_instructions(wb, 'each protein sheet')

    protein_sheets = []
    seen_names = {}

    for fpath in files:
        sheet_name = parse_molecule_name(fpath)

        # Handle duplicate sheet names by appending a counter
        if sheet_name in seen_names:
            seen_names[sheet_name] += 1
            sheet_name = f"{sheet_name[:28]}_{seen_names[sheet_name]}"
        else:
            seen_names[sheet_name] = 0

        print(f"  Parsing: {fpath.name}  →  sheet '{sheet_name}'")
        try:
            records = parse_pdb2pqr(fpath)
            print(f"           {len(records)} residues")
        except ValueError as e:
            print(f"  WARNING: Skipping {fpath.name} — {e}")
            continue

        build_protein_sheet(wb, sheet_name, records, ph_step=ph_step)
        protein_sheets.append(sheet_name)

    if not protein_sheets:
        sys.exit("No valid pKa data found in any file — workbook not saved.")

    # Sheet order: Instructions, all protein sheets alphabetically, AminoAcids
    desired = ['Instructions'] + protein_sheets + ['AminoAcids']
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(wb[name], offset=i - wb.sheetnames.index(name))

    wb.save(output_path)
    print(f"\nSaved:    {output_path}")
    print(f"Sheets:   {', '.join(protein_sheets)}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description="Convert PDB2PQR stderr file(s) into an interactive charge vs pH Excel workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "input",
        nargs="?",
        help="Single PDB2PQR stderr file (single-file mode; omit when using --batch)",
    )
    parser.add_argument(
        "--batch",
        metavar="DIR",
        help="Directory of PDB2PQR stderr files (batch mode)",
    )
    parser.add_argument(
        "--output",
        metavar="FILE",
        help="Output .xlsx path (default: auto-named next to input or inside batch dir)",
    )
    parser.add_argument(
        "--pattern",
        default="*.txt",
        metavar="GLOB",
        help="Glob pattern for input files in batch mode (default: *.txt)",
    )
    parser.add_argument(
        "--template",
        metavar="FILE",
        help="(Single-file mode only) Existing .xlsx to copy other sheets from",
    )
    parser.add_argument(
        "--ph-step",
        type=float,
        default=0.05,
        metavar="STEP",
        help="pH axis step size (default: 0.05, range: 0.05–1.0; use --ph-override for 0.01–0.05)",
    )
    parser.add_argument(
        "--ph-override",
        action="store_true",
        help="Allow --ph-step below 0.05 (down to 0.01). Use with caution: "
             "small steps produce very large files, especially for multi-chain structures.",
    )
    args = parser.parse_args()

    # Validate ph_step
    if args.ph_step > 1.0:
        sys.exit(
            f"ERROR: --ph-step {args.ph_step} is above the maximum of 1.0.\n"
            f"       Valid range is 0.05 to 1.0 (or 0.01 to 1.0 with --ph-override)."
        )
    if args.ph_step < 0.01:
        sys.exit(
            f"ERROR: --ph-step {args.ph_step} is below the absolute minimum of 0.01.\n"
            f"       Valid range is 0.05 to 1.0 (or 0.01 to 1.0 with --ph-override)."
        )
    if args.ph_step < 0.05 and not args.ph_override:
        sys.exit(
            f"WARNING: --ph-step {args.ph_step} is below 0.05, which will produce a very\n"
            f"         large file ({int(round(14.0 / args.ph_step)) + 1} pH rows per chain).\n"
            f"         If you are sure, re-run with both --ph-step {args.ph_step} and --ph-override."
        )

    if args.batch:
        # ── Batch mode ────────────────────────────────────────────────────────
        input_dir = Path(args.batch)
        if not input_dir.is_dir():
            sys.exit(f"Not a directory: {input_dir}")
        output_path = Path(args.output) if args.output else \
                      input_dir / "pKa_Interactive_batch.xlsx"
        batch_main(input_dir, output_path, args.pattern, args.ph_step)

    elif args.input:
        # ── Single-file mode (original behaviour) ─────────────────────────────
        input_path = Path(args.input)
        if not input_path.exists():
            sys.exit(f"File not found: {input_path}")

        sheet_name  = parse_molecule_name(input_path)
        output_path = Path(args.output) if args.output else \
                      input_path.parent / f"{input_path.stem}_pKa_Interactive.xlsx"

        print(f"Parsing:  {input_path}  →  sheet '{sheet_name}'")
        records = parse_pdb2pqr(input_path)
        print(f"Found {len(records)} residues")

        if args.template:
            template = Path(args.template)
            if not template.exists():
                sys.exit(f"Template not found: {template}")
            shutil.copy2(template, output_path)
            wb = openpyxl.load_workbook(output_path)
            print(f"Template: {template}")
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        build_aa_sheet(wb)
        build_instructions(wb, sheet_name)
        build_protein_sheet(wb, sheet_name, records, ph_step=args.ph_step)

        for i, name in enumerate(['Instructions', sheet_name, 'AminoAcids']):
            if name in wb.sheetnames:
                wb.move_sheet(wb[name], offset=i - wb.sheetnames.index(name))

        wb.save(output_path)
        print(f"Saved:    {output_path}")

    else:
        parser.print_help()
        sys.exit(1)
